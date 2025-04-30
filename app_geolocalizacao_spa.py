import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Carregar a base de geolocalizacao padrao (deve estar na mesma pasta do app ou ser incorporada em deploy)
def carregar_base_geolocalizacao():
    path = "NOVA_BASE_DE_GEOLOCALIZACAO_COM_SPA.xlsx"
    df_base = pd.read_excel(path)
    df_base.rename(columns={"SP": "Rodovia"}, inplace=True)
    df_base['Rodovia'] = df_base['Rodovia'].apply(ajustar_rodovia)
    df_base = df_base[df_base['Unnamed: 7'].str.contains(",", na=False)].copy()
    df_base[['x', 'y']] = df_base['Unnamed: 7'].str.split(',', expand=True)[[0, 1]].astype(float)
    return df_base

def ajustar_rodovia(rod):
    if isinstance(rod, str) and rod.startswith("SP "):
        try:
            num = int(rod.split(" ")[1])
            return f"SP {num:03d}"
        except:
            return rod
    return rod

def preencher_coordenadas(df_ocorrencias, df_base):
    df_ocorrencias['Rodovia'] = df_ocorrencias['Rodovia'].apply(ajustar_rodovia)
    df_ocorrencias['x'] = df_ocorrencias.get('x', pd.NA)
    df_ocorrencias['y'] = df_ocorrencias.get('y', pd.NA)
    df_ocorrencias['preenchido'] = False

    for idx, row in df_ocorrencias.iterrows():
        rod = row['Rodovia']
        km = row['km']
        if pd.notnull(km):
            candidatos = df_base[(df_base['Rodovia'] == rod) & (abs(df_base['km'] - km) <= 0.5)]
            if not candidatos.empty:
                df_ocorrencias.at[idx, 'x'] = candidatos.iloc[0]['x']
                df_ocorrencias.at[idx, 'y'] = candidatos.iloc[0]['y']
                df_ocorrencias.at[idx, 'preenchido'] = True
    return df_ocorrencias

def gerar_excel_colorido(df):
    output = BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    header = {cell.value: i for i, cell in enumerate(ws[1])}
    x_col = header.get("x", None)
    y_col = header.get("y", None)
    p_col = header.get("preenchido", None)
    fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    if None not in (x_col, y_col, p_col):
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, p_col + 1).value:
                ws.cell(row, x_col + 1).fill = fill
                ws.cell(row, y_col + 1).fill = fill

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output

# Interface Streamlit
st.title("Preenchimento de Coordenadas - Rodovias SP")
st.write("Esse app preenche as coordenadas x e y com base na rodovia e km (Â±0,5), utilizando uma base interna de geolocalizaÃ§Ã£o.")

st.markdown("""
**ðŸ“ OrientaÃ§Ã£o:** Para a localizaÃ§Ã£o do lat/long Ã© necessÃ¡rio que o arquivo tenha duas colunas:
- **\"Rodovia\"** com o nome igual Ã  formataÃ§Ã£o do DER. Por exemplo: "SP 008", "SP 058", "SPA 099/060".
- **\"km\"** com o valor do quilÃ´metro em nÃºmero inteiro ou decimal (separado por vÃ­rgula).

Ã‰ importante que os tÃ­tulos (rÃ³tulos) da coluna estejam **exatamente iguais** ao exemplo acima.
""")

uploaded_ocorrencias = st.file_uploader("ðŸ“Ž Planilha de ocorrÃªncias (xlsx com 'Rodovia' e 'km')", type="xlsx")

if uploaded_ocorrencias:
    df_base = carregar_base_geolocalizacao()
    df_ocorrencias = pd.read_excel(uploaded_ocorrencias)

    if st.button("ðŸš€ Processar e Preencher Coordenadas"):
        df_resultado = preencher_coordenadas(df_ocorrencias.copy(), df_base)
        st.success(f"Preenchimento concluÃ­do! {df_resultado['preenchido'].sum()} coordenadas preenchidas.")

        st.dataframe(df_resultado.head(20))

        excel_result = gerar_excel_colorido(df_resultado)
        st.download_button(
            label="ðŸ“¥ Baixar planilha com preenchimentos em vermelho",
            data=excel_result,
            file_name="ocorrencias_com_coordenadas.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
